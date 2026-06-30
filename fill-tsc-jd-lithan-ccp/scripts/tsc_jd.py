#!/usr/bin/env python3
"""Read or write the JD (Job responsibilities) column of a Lithan CCP TSC Scoping Matrix sheet.

The TSC matrix has one sheet per job role. Each sheet has a competency table whose
header row contains "Technical Skills & Competencies". Columns are located by header
text (positions are stable but row numbers vary between sheets), so this script finds
them dynamically rather than hardcoding cell addresses.

Usage:
  # List sheets in the workbook
  python3 tsc_jd.py sheets "<file.xlsx>"

  # Dump each skill row of a sheet as JSON: row, skill, proficiency, objective,
  # include flag (E), current tools (G), and current JD (F). Use this to read the
  # objectives you write the JD from, and to copy the style of an already-filled sheet.
  python3 tsc_jd.py read "<file.xlsx>" "<sheet name>"

  # Write JD text into column F from a JSON map {row: text}. Formatting is copied
  # from the Objectives (D) cell of the same row so font/wrap/alignment match.
  # Recalculates with LibreOffice afterwards and reports any formula errors.
  python3 tsc_jd.py write "<file.xlsx>" "<sheet name>" '<json {"10":"...","11":"..."}>'
  python3 tsc_jd.py write "<file.xlsx>" "<sheet name>" @path/to/map.json
"""
import sys, json, copy, subprocess, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HEADER_KEY = "technical skills"   # lowercased substring identifying the header row
COLS = {"skill": "B", "prof": "C", "obj": "D", "include": "E",
        "jd": "F", "tools": "G", "hours": "H", "mode": "I"}
COLNUM = {k: ord(v) - 64 for k, v in COLS.items()}


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        v = ws.cell(r, COLNUM["skill"]).value
        if v and HEADER_KEY in str(v).lower():
            return r
    raise SystemExit("Could not find header row (no 'Technical Skills' in column B).")


def skill_rows(ws, header_row):
    """Skill data rows: below the header (and its sub-label row), stop at the
    'Mentor Assessment' section or first fully blank skill/objective pair."""
    rows = []
    r = header_row + 1
    # skip the sub-label row that has '(aligned with JD)' under F
    while r <= ws.max_row:
        b = ws.cell(r, COLNUM["skill"]).value
        if b and "mentor assessment" in str(b).lower():
            break
        bt = str(b).strip() if b else ""
        # a real skill row has a skill name in B and an objective in D
        d = ws.cell(r, COLNUM["obj"]).value
        sub = bt == "" and not d
        is_sublabel = ws.cell(r, COLNUM["jd"]).value and "aligned with jd" in str(ws.cell(r, COLNUM["jd"]).value).lower()
        if bt and d and not is_sublabel:
            rows.append(r)
        r += 1
    return rows


def cmd_sheets(path):
    wb = load_workbook(path, read_only=True)
    print(json.dumps(wb.sheetnames, indent=2))


def cmd_read(path, sheet):
    wb = load_workbook(path)
    ws = wb[sheet]
    hr = find_header_row(ws)
    out = []
    for r in skill_rows(ws, hr):
        out.append({
            "row": r,
            "skill": _s(ws.cell(r, COLNUM["skill"]).value),
            "proficiency": _s(ws.cell(r, COLNUM["prof"]).value),
            "include": _s(ws.cell(r, COLNUM["include"]).value),
            "objective": _s(ws.cell(r, COLNUM["obj"]).value),
            "tools": _s(ws.cell(r, COLNUM["tools"]).value),
            "jd_current": _s(ws.cell(r, COLNUM["jd"]).value),
        })
    print(json.dumps({"sheet": sheet, "header_row": hr, "rows": out}, indent=2, ensure_ascii=False))


def cmd_write(path, sheet, payload):
    if payload.startswith("@"):
        with open(payload[1:], encoding="utf-8") as f:
            mapping = json.load(f)
    else:
        mapping = json.loads(payload)
    wb = load_workbook(path)
    ws = wb[sheet]
    jd_c = COLNUM["jd"]
    obj_c = COLNUM["obj"]
    n = 0
    for k, text in mapping.items():
        r = int(k)
        cell = ws.cell(r, jd_c)
        cell.value = text
        ref = ws.cell(r, obj_c)          # match the Objectives cell styling
        cell.font = copy.copy(ref.font)
        cell.alignment = copy.copy(ref.alignment)
        n += 1
    wb.save(path)
    print(f"Wrote {n} JD cells to column F of '{sheet}'.")
    _recalc(path)


def _recalc(path):
    recalc = os.path.join(os.path.dirname(__file__), "..", "..", "xlsx", "scripts", "recalc.py")
    recalc = os.path.normpath(recalc)
    if os.path.exists(recalc):
        try:
            out = subprocess.run([sys.executable, recalc, path], capture_output=True, text=True, timeout=120)
            tail = out.stdout.strip().splitlines()[-6:]
            print("recalc:", " ".join(tail))
        except Exception as e:
            print(f"(recalc skipped: {e})")
    else:
        print("(recalc.py from xlsx skill not found; open in Excel to recalc if needed)")


def _s(v):
    return None if v is None else str(v).strip()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == "sheets":
        cmd_sheets(sys.argv[2])
    elif cmd == "read":
        cmd_read(sys.argv[2], sys.argv[3])
    elif cmd == "write":
        cmd_write(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        print(__doc__); sys.exit(1)
